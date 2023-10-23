import openpyxl

wb = openpyxl.load_workbook('transactions.xlsx')
ws = wb.active

ws['D1'] = 'corrected'


for i in range(2, ws.max_row + 1):
    price = ws[f'C{i}'].value
    price_float = float(price.replace('€', '').replace(',', '.'))

    discount_price = round(price_float * 0.9, 2)

    ws[f'D{i}'].value = discount_price

    wb.save('transactions_discount.xlsx')
